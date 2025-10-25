import streamlit as st
import pandas as pd
import pykakasi
import xlrd  # .xls ファイル用
import openpyxl  # .xlsx ファイル用
from openpyxl import load_workbook
from io import BytesIO

def convert_to_hiragana(text):
    """テキストをひらがなに変換する関数"""
    if pd.isna(text):
        return ""
    if not isinstance(text, str):
        text = str(text)
    
    kks = pykakasi.kakasi()
    result = kks.convert(text)
    hiragana = ''.join([item['hira'] for item in result])
    return hiragana

def process_excel_file(file):
    """Excelファイルの全シートの全セルの内容をひらがなに変換する"""
    # ファイル拡張子の確認
    file_extension = file.name.split('.')[-1].lower()

    # ストリームをバイトに読み込み、再利用できる BytesIO を作る
    file_bytes = file.read()
    file_io = BytesIO(file_bytes)

    # 各シートのDataFrameを保存する辞書
    dfs = {}
    output_bytes = None

    # .xlsx の場合は openpyxl でワークブックを開き、セルの値を直接書き換えてスタイルを保持
    if file_extension == 'xlsx':
        wb = load_workbook(filename=BytesIO(file_bytes))

        for ws in wb.worksheets:
            # 表示用のDataFrameは pandas で読み込んでおく（型処理が楽なため）
            df = pd.read_excel(BytesIO(file_bytes), sheet_name=ws.title)

            # ワークシート上のセルを走査して文字列セルのみ変換（数値や日時、数式は保持）
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    # 文字列のみ変換（NaNやNone、数値はそのまま）
                    if isinstance(val, str):
                        cell.value = convert_to_hiragana(val)

            # DataFrame側は従来通り列ごとに変換して表示用に使う
            for column in df.columns:
                df[column] = df[column].apply(convert_to_hiragana)

            dfs[ws.title] = df

        # 変更済みのワークブックをバイト列に保存（書式・スタイルが保持される）
        out = BytesIO()
        wb.save(out)
        output_bytes = out.getvalue()

    else:
        # .xls やその他は pandas + xlrd（フォーマット情報は維持できない）
        excel_file = pd.ExcelFile(BytesIO(file_bytes), engine='xlrd')
        sheet_names = excel_file.sheet_names

        out = BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            for sheet_name in sheet_names:
                df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, engine='xlrd')
                for column in df.columns:
                    df[column] = df[column].apply(convert_to_hiragana)
                dfs[sheet_name] = df
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        output_bytes = out.getvalue()

    return dfs, output_bytes

def main():
    st.title("Excel ひらがな変換アプリ")
    st.write("Excelファイルのすべてのセルをひらがなに変換します")
    
    # ファイルアップローダーの表示
    uploaded_file = st.file_uploader("Excelファイルを選択してください", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        try:
            # ファイルの処理（DataFrame辞書と出力バイト列を受け取る）
            converted_dfs, output_bytes = process_excel_file(uploaded_file)

            # 変換結果の表示
            st.write("変換結果:")
            for sheet_name, df in converted_dfs.items():
                st.write(f"シート: {sheet_name}")
                st.dataframe(df)
                st.divider()  # シート間の区切り線

            # .xlsx の場合は書式を引き継いでいる旨を表示
            ext = uploaded_file.name.split('.')[-1].lower()
            if ext == 'xlsx':
                st.info(".xlsx の書式（フォントやセル背景など）は保持されています。")
            else:
                st.warning(".xls ファイルは書式情報を維持できないため、変換後のダウンロードでは書式が失われる可能性があります。可能であれば .xlsx に変換してからアップロードしてください。")

            # ダウンロードボタンの表示（既に生成したバイト列を直接提供）
            # 元のファイル名を元にダウンロード名を作成（例: 元が report.xlsx -> report_ひらがな.xlsx）
            original_base = uploaded_file.name.rsplit('.', 1)[0]
            download_name = f"{original_base}_ひらがな.xlsx"
            st.download_button(
                label="変換済みファイルをダウンロード",
                data=output_bytes,
                file_name=download_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"エラーが発生しました: {str(e)}")

if __name__ == "__main__":
    main()
